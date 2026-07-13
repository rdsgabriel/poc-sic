"""
Gera as planilhas PGR (riscos) e PCMSO (exames) a partir dos GHEs extraídos.

Formato conforme PROMPT_PRODUCAO_PLANILHAS_PCMSO.md e templates de importação:
- 1 linha por função/cargo de cada GHE
- Setor no padrão "GHE 01 - ADMINISTRATIVO"
- listas separadas por vírgula SEM espaço
- Periódico no formato "Exame,MESES,Exame2,MESES2"
"""

from __future__ import annotations

import re
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .extractors.base import GHE

PGR_HEADERS = [
    "Setor", "Função", "Função Inativa",
    "Químico", "Físico", "Biológico", "Ergonômicos",
    "Riscos do Tipo Ergonomicos Biomecanicos",
    "Riscos do Tipo Ergonômicos – Mobiliario e Equipamentos",
    "Riscos do Tipo Ergonômicos – Organizacionais",
    "Riscos do Tipo Ergonômicos – Psicossociais",
    "Riscos do Tipo Ergonômicos – Ambientais",
    "Acidentes", "Riscos do Tipo Periculosos", "Riscos do Tipo Penosos",
    "Risco com Associação de Riscos", "Riscos para Ausência de Risco",
    "Outros Tipos",
    "NR_10 (Eletricidade)", "NR 11 (Operação de Veículos Industriais)",
    "NR_20 (Segurança e Saúde no Trabalho com Inflamáveis e Combustíveis)",
    "NR_30 (Trabalho Offshore)", "NR_33 (Espaço Confinado)",
    "NR_34 (Apto para Brigada de Emergência)", "NR_35 (Trabalho em Altura)",
    "NR_37 (Apto para Efetuar Transbordo)", "NR_37 (Embarque Offshore)",
    "NR_37 (Trabalho em Brigada de Emergência)", "DESCRICAO",
]

PCMSO_HEADERS = [
    "Setor", "Função",
    "Perfil Admissional", "Perfil Periódico", "Perfil Demissional",
    "Perfil Mudança de função", "Perfil Periódico Semestral",
    "Perfil Retorno ao Trabalho", "Perfil Semestral Após Admissão",
]

GRUPO_COL = {  # grupo de risco -> índice (0-based) da coluna na PGR
    "Químico": 3, "Físico": 4, "Biológico": 5, "Ergonômicos": 6,
    "Acidente": 12, "Periculoso": 13, "Penoso": 14,
}

# risco (normalizado, por substring) -> coluna NR
NR_TRIGGERS = [
    ("choque eletrico", 18),                        # NR_10
    ("espaco confinado", 22),                       # NR_33
    ("queda de diferenca", 24),                     # NR_35 (nível >= 2m)
    ("trabalho em altura", 24),                     # NR_35 (nomenclatura VIX)
]
NR_MARK = "X"


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def _estilo_header(ws, n_cols: int) -> None:
    fill = PatternFill("solid", fgColor="D9E1F2")
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = fill


def montar_pgr(ghes: list[GHE]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "PGR"
    ws.append(PGR_HEADERS)
    _estilo_header(ws, len(PGR_HEADERS))

    for ghe in ghes:
        linha_base = [""] * len(PGR_HEADERS)
        linha_base[0] = ghe.setor

        por_grupo: dict[int, list[str]] = {}
        for r in ghe.riscos:
            col = GRUPO_COL.get(r.grupo)
            if col is not None:
                por_grupo.setdefault(col, []).append(r.nome)
        for col, nomes in por_grupo.items():
            linha_base[col] = ",".join(nomes)

        if not ghe.riscos:
            linha_base[16] = "Ausência de Riscos"

        riscos_norm = [_norm(r.nome) for r in ghe.riscos]
        for trigger, col in NR_TRIGGERS:
            if any(trigger in rn for rn in riscos_norm):
                linha_base[col] = NR_MARK

        for cargo in ghe.cargos:
            linha = list(linha_base)
            linha[1] = cargo
            ws.append(linha)
    return wb


def montar_pcmso(ghes: list[GHE]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "PCMSO"
    ws.append(PCMSO_HEADERS)
    _estilo_header(ws, len(PCMSO_HEADERS))

    for ghe in ghes:
        admissional = [e.nome for e in ghe.exames if e.admissao]
        demissional = [e.nome for e in ghe.exames if e.demissao]
        mudanca = [e.nome for e in ghe.exames if e.mud_riscos]
        retorno = [e.nome for e in ghe.exames if e.ret_trab]

        periodico: list[str] = []
        periodico_sem: list[str] = []
        for e in ghe.exames:
            if e.periodico_meses:
                if e.periodico_meses == 6:
                    periodico_sem.append(e.nome)
                else:
                    periodico.extend([e.nome, str(e.periodico_meses)])

        semestral_apos = [
            e.nome for e in ghe.exames if e.apos_adm or e.apos_adm_meses
        ]

        linha_base = [
            ghe.setor, "",
            ",".join(admissional),
            ",".join(periodico),
            ",".join(demissional),
            ",".join(mudanca),
            ",".join(periodico_sem),
            ",".join(retorno),
            ",".join(semestral_apos),
        ]
        for cargo in ghe.cargos:
            linha = list(linha_base)
            linha[1] = cargo
            ws.append(linha)
    return wb
